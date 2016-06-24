#!/usr/bin/python
# -*- coding: utf-8 -*-
# pttrnsrch.py

import sys
import re
import os
import openpyxl



# args validation
if ((len(sys.argv) != 3) or (sys.argv[1] not in ["-e", "-n"]) or (not os.path.isdir(sys.argv[2]))):
	sys.exit("wrong args, try:\n./pttrnsrch.py -e|-n pathname")
else:
	action = sys.argv[1]
	path = sys.argv[2]


# define pattern
if (action == "-e"):
	pattern = r'EXONERAR(, a pedido,)? (.*?) do cargo (.*?), código (DAS ?-? ?\d\d\d\.\d)'
else:
	pattern = r'NOMEAR'
regexp = re.compile(pattern)


# files in path
try:
	files = []
	content = os.listdir(path)
	content.sort()
	for c in content:
		if not os.path.isdir(path + c):
			files.append(c)
except OSError as e:
	print "OSError({0}): {1}".format(e.errno, e.strerror)


# read file content and replace line breaks
try:
	wb = openpyxl.Workbook()
	ws = wb.active
	ws['A1'] = "Pedido ou nao"
	ws['B1'] = "Nome"
	ws['C1'] = "Cargo/Orgao"
	ws['D1'] = "Codigo"
	index = 2
	for f in files:
		fo = open(path + f, 'r')
		txt = fo.read()
		txt = re.sub(r'\n', ' ', txt)
		txt = re.sub(r' {2,}', ' ', txt)
		results = regexp.findall(txt)
		for m in results:
			line = []
			for n in m:
				n = n.decode('utf-8')
				if (n == ", a pedido,"):
					line.append("a pedido")
				elif not n:
					line.append("nao pedido")
				else:
					line.append(n)
			ws['A'+str(index)] = line[0]
			ws['B'+str(index)] = line[1]
			ws['C'+str(index)] = line[2]
			ws['D'+str(index)] = line[3]
			print line
			index += 1
	wb.save('dou_search.xlsx')
	print index
except IOError:
	print "cannot find or read the data from file", f
	if file_is_open(fo):
		fo.close()



print "\nEOF"
